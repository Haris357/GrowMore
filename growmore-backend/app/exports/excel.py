"""Excel Export Generator using openpyxl."""

import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """Generate Excel exports."""

    def __init__(self):
        # Style definitions
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        self.body_font = Font(size=10)
        self.number_alignment = Alignment(horizontal="right")
        self.text_alignment = Alignment(horizontal="left")

        self.border = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB"),
        )

        self.alt_row_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        self.positive_font = Font(color="059669", size=10)
        self.negative_font = Font(color="DC2626", size=10)

    def _style_header(self, ws, row: int, columns: int):
        """Apply header styling to a row."""
        for col in range(1, columns + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

    def _style_body(self, ws, start_row: int, end_row: int, columns: int):
        """Apply body styling to rows."""
        for row in range(start_row, end_row + 1):
            for col in range(1, columns + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = self.body_font
                cell.border = self.border
                if (row - start_row) % 2 == 1:
                    cell.fill = self.alt_row_fill

    def _auto_column_width(self, ws, columns: int):
        """Auto-adjust column widths."""
        for col in range(1, columns + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def generate_portfolio_excel(
        self,
        user_name: str,
        report_data: Dict[str, Any],
    ) -> bytes:
        """
        Generate portfolio report Excel.

        Args:
            user_name: User's name
            report_data: Portfolio analytics data

        Returns:
            Excel file bytes
        """
        wb = Workbook()

        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        ws_summary["A1"] = "Portfolio Report"
        ws_summary["A1"].font = Font(bold=True, size=16)
        ws_summary["A2"] = f"User: {user_name}"
        ws_summary["A3"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"

        summary = report_data.get("summary", {})
        summary_start = 5

        summary_data = [
            ["Metric", "Value"],
            ["Total Value", f"PKR {summary.get('total_value', 0):,.2f}"],
            ["Total Invested", f"PKR {summary.get('total_invested', 0):,.2f}"],
            ["Total Gain/Loss", f"PKR {summary.get('total_gain_loss', 0):,.2f}"],
            ["Return %", f"{summary.get('gain_loss_percentage', 0):.2f}%"],
            ["Holdings Count", summary.get('holdings_count', 0)],
        ]

        for row_idx, row_data in enumerate(summary_data):
            for col_idx, value in enumerate(row_data):
                ws_summary.cell(row=summary_start + row_idx, column=col_idx + 1, value=value)

        self._style_header(ws_summary, summary_start, 2)
        self._style_body(ws_summary, summary_start + 1, summary_start + len(summary_data) - 1, 2)
        self._auto_column_width(ws_summary, 2)

        # Holdings Sheet
        holdings = report_data.get("holdings", [])
        if holdings:
            ws_holdings = wb.create_sheet("Holdings")

            headers = ["Symbol", "Name", "Sector", "Qty", "Avg Price", "Current",
                      "Invested", "Value", "Gain/Loss", "Gain %", "Day %"]
            for col, header in enumerate(headers, 1):
                ws_holdings.cell(row=1, column=col, value=header)

            for row_idx, h in enumerate(holdings, 2):
                ws_holdings.cell(row=row_idx, column=1, value=h.get("symbol", ""))
                ws_holdings.cell(row=row_idx, column=2, value=h.get("name", ""))
                ws_holdings.cell(row=row_idx, column=3, value=h.get("sector", ""))
                ws_holdings.cell(row=row_idx, column=4, value=h.get("quantity", 0))
                ws_holdings.cell(row=row_idx, column=5, value=h.get("average_price", 0))
                ws_holdings.cell(row=row_idx, column=6, value=h.get("current_price", 0))
                ws_holdings.cell(row=row_idx, column=7, value=h.get("total_invested", 0))
                ws_holdings.cell(row=row_idx, column=8, value=h.get("current_value", 0))
                ws_holdings.cell(row=row_idx, column=9, value=h.get("gain_loss", 0))

                gain_pct = h.get("gain_loss_pct", 0)
                cell = ws_holdings.cell(row=row_idx, column=10, value=f"{gain_pct:.2f}%")
                cell.font = self.positive_font if gain_pct >= 0 else self.negative_font

                day_pct = h.get("day_change_pct", 0)
                cell = ws_holdings.cell(row=row_idx, column=11, value=f"{day_pct:.2f}%")
                cell.font = self.positive_font if day_pct >= 0 else self.negative_font

            self._style_header(ws_holdings, 1, len(headers))
            self._style_body(ws_holdings, 2, len(holdings) + 1, len(headers))
            self._auto_column_width(ws_holdings, len(headers))

        # Sector Breakdown Sheet
        sectors = report_data.get("sector_breakdown", [])
        if sectors:
            ws_sectors = wb.create_sheet("Sectors")

            headers = ["Sector", "Value (PKR)", "Allocation %", "Holdings"]
            for col, header in enumerate(headers, 1):
                ws_sectors.cell(row=1, column=col, value=header)

            for row_idx, s in enumerate(sectors, 2):
                ws_sectors.cell(row=row_idx, column=1, value=s.get("sector", ""))
                ws_sectors.cell(row=row_idx, column=2, value=s.get("value", 0))
                ws_sectors.cell(row=row_idx, column=3, value=f"{s.get('percentage', 0):.2f}%")
                ws_sectors.cell(row=row_idx, column=4, value=s.get("holdings_count", 0))

            self._style_header(ws_sectors, 1, len(headers))
            self._style_body(ws_sectors, 2, len(sectors) + 1, len(headers))
            self._auto_column_width(ws_sectors, len(headers))

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_transactions_excel(
        self,
        transactions: List[Dict[str, Any]],
        user_name: Optional[str] = None,
    ) -> bytes:
        """Generate transactions Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Title
        if user_name:
            ws["A1"] = f"Transaction History - {user_name}"
            ws["A1"].font = Font(bold=True, size=14)
            start_row = 3
        else:
            start_row = 1

        headers = ["Date", "Type", "Symbol", "Quantity", "Price", "Total", "Fees", "Status"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=start_row, column=col, value=header)

        for row_idx, t in enumerate(transactions, start_row + 1):
            date = t.get("date", "")
            if isinstance(date, datetime):
                date = date.strftime("%Y-%m-%d %H:%M")

            ws.cell(row=row_idx, column=1, value=date)
            ws.cell(row=row_idx, column=2, value=t.get("type", "").upper())
            ws.cell(row=row_idx, column=3, value=t.get("symbol", ""))
            ws.cell(row=row_idx, column=4, value=t.get("quantity", 0))
            ws.cell(row=row_idx, column=5, value=t.get("price", 0))
            ws.cell(row=row_idx, column=6, value=t.get("total", 0))
            ws.cell(row=row_idx, column=7, value=t.get("fees", 0))
            ws.cell(row=row_idx, column=8, value=t.get("status", ""))

        self._style_header(ws, start_row, len(headers))
        self._style_body(ws, start_row + 1, start_row + len(transactions), len(headers))
        self._auto_column_width(ws, len(headers))

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_watchlist_excel(
        self,
        watchlist_name: str,
        stocks: List[Dict[str, Any]],
    ) -> bytes:
        """Generate watchlist Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = watchlist_name[:31]  # Excel sheet name limit

        ws["A1"] = f"Watchlist: {watchlist_name}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"

        start_row = 4
        headers = ["Symbol", "Name", "Price", "Change", "Change %", "Volume", "Market Cap"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=start_row, column=col, value=header)

        for row_idx, s in enumerate(stocks, start_row + 1):
            ws.cell(row=row_idx, column=1, value=s.get("symbol", ""))
            ws.cell(row=row_idx, column=2, value=s.get("name", ""))
            ws.cell(row=row_idx, column=3, value=s.get("current_price", 0))

            change = s.get("change_amount", 0) or 0
            cell = ws.cell(row=row_idx, column=4, value=change)
            cell.font = self.positive_font if change >= 0 else self.negative_font

            change_pct = s.get("change_percentage", 0) or 0
            cell = ws.cell(row=row_idx, column=5, value=f"{change_pct:.2f}%")
            cell.font = self.positive_font if change_pct >= 0 else self.negative_font

            ws.cell(row=row_idx, column=6, value=s.get("volume", 0))
            ws.cell(row=row_idx, column=7, value=s.get("market_cap", 0))

        self._style_header(ws, start_row, len(headers))
        self._style_body(ws, start_row + 1, start_row + len(stocks), len(headers))
        self._auto_column_width(ws, len(headers))

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_generic_excel(
        self,
        data: List[Dict[str, Any]],
        sheet_name: str = "Data",
        title: Optional[str] = None,
        columns: Optional[List[str]] = None,
    ) -> bytes:
        """
        Generate generic Excel from list of dictionaries.

        Args:
            data: List of dictionaries
            sheet_name: Name for the sheet
            title: Optional title
            columns: Optional list of column keys

        Returns:
            Excel file bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        start_row = 1
        if title:
            ws["A1"] = title
            ws["A1"].font = Font(bold=True, size=14)
            start_row = 3

        if not data:
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()

        # Get columns
        if columns is None:
            columns = list(data[0].keys())

        # Write headers
        headers = [col.replace("_", " ").title() for col in columns]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=start_row, column=col_idx, value=header)

        # Write data
        for row_idx, row_data in enumerate(data, start_row + 1):
            for col_idx, col_key in enumerate(columns, 1):
                value = row_data.get(col_key, "")
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M")
                elif isinstance(value, (dict, list)):
                    value = str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        self._style_header(ws, start_row, len(headers))
        self._style_body(ws, start_row + 1, start_row + len(data), len(headers))
        self._auto_column_width(ws, len(headers))

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
